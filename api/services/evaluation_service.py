import csv
import io
import json
import logging
from collections.abc import Mapping
from typing import Any, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from configs import dify_config
from core.evaluation.entities.evaluation_entity import (
    METRIC_NODE_TYPE_MAPPING,
    METRIC_VALUE_TYPE_MAPPING,
    DefaultMetric,
    EvaluationCategory,
    EvaluationConfigData,
    EvaluationDatasetInput,
    EvaluationItemResult,
    EvaluationMetric,
    EvaluationMetricName,
    EvaluationRunData,
    EvaluationRunRequest,
    NodeInfo,
)
from core.evaluation.entities.judgment_entity import JudgmentConfig, JudgmentResult
from core.evaluation.evaluation_manager import EvaluationManager
from core.evaluation.judgment.processor import JudgmentProcessor
from graphon.enums import WorkflowNodeExecutionMetadataKey
from graphon.node_events.base import NodeRunResult
from libs.datetime_utils import naive_utc_now
from models.evaluation import (
    EvaluationConfiguration,
    EvaluationRun,
    EvaluationRunItem,
    EvaluationRunStatus,
    EvaluationTargetType,
)
from models.model import App, AppMode
from models.snippet import CustomizedSnippet
from models.workflow import Workflow
from services.errors.evaluation import (
    EvaluationDatasetInvalidError,
    EvaluationFrameworkNotConfiguredError,
    EvaluationMaxConcurrentRunsError,
    EvaluationNotFoundError,
)
from services.snippet_service import SnippetService
from services.workflow_service import WorkflowService

logger = logging.getLogger(__name__)


class EvaluationService:
    """
    Service for evaluation-related operations.

    Provides functionality to generate evaluation dataset templates
    based on App or Snippet input parameters.
    """

    # Excluded app modes that don't support evaluation templates
    EXCLUDED_APP_MODES = {AppMode.RAG_PIPELINE}
    CONSOLE_DISABLED_CATEGORIES = {EvaluationCategory.AGENT}
    CONSOLE_DISABLED_METRICS = {
        EvaluationMetricName.TOOL_CORRECTNESS.value,
        EvaluationMetricName.TASK_COMPLETION.value,
    }
    METRICS_REQUIRING_EXPECTED_OUTPUT = {
        EvaluationMetricName.ANSWER_CORRECTNESS.value,
        EvaluationMetricName.SEMANTIC_SIMILARITY.value,
        EvaluationMetricName.CONTEXT_PRECISION.value,
        EvaluationMetricName.CONTEXT_RECALL.value,
    }

    @classmethod
    def generate_dataset_template(
        cls,
        target: Union[App, CustomizedSnippet],
        target_type: str,
    ) -> tuple[bytes, str]:
        """
        Generate evaluation dataset template as XLSX bytes.

        Creates an XLSX file with headers based on the evaluation target's input parameters.
        The first column is index, followed by input parameter columns.

        :param target: App or CustomizedSnippet instance
        :param target_type: Target type string ("apps" or "snippets")
        :return: Tuple of (xlsx_content_bytes, filename)
        :raises ValueError: If target type is not supported or app mode is excluded
        """
        # Validate target type
        if target_type == EvaluationTargetType.APPS.value:
            if not isinstance(target, App):
                raise ValueError("Invalid target: expected App instance")
            if AppMode.value_of(target.mode) in cls.EXCLUDED_APP_MODES:
                raise ValueError(f"App mode '{target.mode}' does not support evaluation templates")
            input_fields = cls._get_app_input_fields(target)
        elif target_type == EvaluationTargetType.SNIPPETS.value:
            if not isinstance(target, CustomizedSnippet):
                raise ValueError("Invalid target: expected CustomizedSnippet instance")
            input_fields = cls._get_snippet_input_fields(target)
        else:
            raise ValueError(f"Unsupported target type: {target_type}")

        # Generate XLSX template
        xlsx_content = cls._generate_xlsx_template(input_fields, target.name)

        # Build filename
        truncated_name = target.name[:10] + "..." if len(target.name) > 10 else target.name
        filename = f"{truncated_name}-evaluation-dataset.xlsx"

        return xlsx_content, filename

    @classmethod
    def _get_app_input_fields(cls, app: App) -> list[dict]:
        """
        Get input fields from App's workflow.

        :param app: App instance
        :return: List of input field definitions
        """
        workflow_service = WorkflowService()
        workflow = workflow_service.get_published_workflow(app_model=app)
        if not workflow:
            workflow = workflow_service.get_draft_workflow(app_model=app)

        if not workflow:
            return []

        # Get user input form from workflow
        user_input_form = workflow.user_input_form()
        return user_input_form

    @classmethod
    def _get_snippet_input_fields(cls, snippet: CustomizedSnippet) -> list[dict]:
        """
        Get input fields from Snippet.

        Tries to get from snippet's own input_fields first,
        then falls back to workflow's user_input_form.

        :param snippet: CustomizedSnippet instance
        :return: List of input field definitions
        """
        # Try snippet's own input_fields first
        input_fields = snippet.input_fields_list
        if input_fields:
            return input_fields

        # Fallback to workflow's user_input_form
        snippet_service = SnippetService()
        workflow = snippet_service.get_published_workflow(snippet=snippet)
        if not workflow:
            workflow = snippet_service.get_draft_workflow(snippet=snippet)

        if workflow:
            return workflow.user_input_form()

        return []

    @classmethod
    def _generate_xlsx_template(cls, input_fields: list[dict], target_name: str) -> bytes:
        """
        Generate XLSX template file content.

        Creates a workbook with:
        - First row as header row with "index" and input field names
        - Styled header with background color and borders
        - Empty data rows ready for user input

        :param input_fields: List of input field definitions
        :param target_name: Name of the target (for sheet name)
        :return: XLSX file content as bytes
        """
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Evaluation Dataset")

        sheet_name = "Evaluation Dataset"
        ws.title = sheet_name

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Build header row
        headers = ["index"]

        for field in input_fields:
            field_label = str(field.get("label") or field.get("variable") or "")
            headers.append(field_label)

        # Write header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Set column widths
        ws.column_dimensions["A"].width = 10  # index column
        for col_idx in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        # Add one empty row with row number for user reference
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col_idx, value="")
            cell.border = thin_border
            if col_idx == 1:
                cell.value = 1
                cell.alignment = Alignment(horizontal="center")

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    @classmethod
    def generate_retrieval_dataset_template(cls) -> tuple[bytes, str]:
        """Generate evaluation dataset XLSX template for knowledge base retrieval.

        The template contains three columns: ``index``, ``query``, and
        ``expected_output``.  Callers upload a filled copy and start an
        evaluation run with ``target_type="dataset"``.

        :returns: (xlsx_content_bytes, filename)
        """
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Evaluation Dataset")
        ws.title = "Evaluation Dataset"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = ["index", "query", "expected_output"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 30

        # Add one sample row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col_idx, value="")
            cell.border = thin_border
            if col_idx == 1:
                cell.value = 1
                cell.alignment = Alignment(horizontal="center")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue(), "retrieval-evaluation-dataset.xlsx"

    # ---- Evaluation Configuration CRUD ----

    @classmethod
    def get_evaluation_config(
        cls,
        session: Session,
        tenant_id: str,
        target_type: str,
        target_id: str,
    ) -> EvaluationConfiguration | None:
        return (
            session.query(EvaluationConfiguration)
            .filter_by(tenant_id=tenant_id, target_type=target_type, target_id=target_id)
            .first()
        )

    @classmethod
    def save_evaluation_config(
        cls,
        session: Session,
        tenant_id: str,
        target_type: str,
        target_id: str,
        account_id: str,
        data: EvaluationConfigData,
    ) -> EvaluationConfiguration:
        config = cls.get_evaluation_config(session, tenant_id, target_type, target_id)
        if config is None:
            config = EvaluationConfiguration(
                tenant_id=tenant_id,
                target_type=target_type,
                target_id=target_id,
                created_by=account_id,
                updated_by=account_id,
            )
            session.add(config)

        config.evaluation_model_provider = data.evaluation_model_provider
        config.evaluation_model = data.evaluation_model
        config.metrics_config = json.dumps(
            {
                "default_metrics": [m.model_dump() for m in data.default_metrics],
                "customized_metrics": data.customized_metrics.model_dump() if data.customized_metrics else None,
            }
        )
        config.judgement_conditions = json.dumps(data.judgment_config.model_dump() if data.judgment_config else {})
        config.customized_workflow_id = (
            data.customized_metrics.evaluation_workflow_id if data.customized_metrics else None
        )
        config.updated_by = account_id
        session.commit()
        session.refresh(config)
        return config

    @classmethod
    def list_targets_by_customized_workflow(
        cls,
        session: Session,
        tenant_id: str,
        customized_workflow_id: str,
    ) -> list[EvaluationConfiguration]:
        """Return all evaluation configs that reference the given workflow as customized metrics."""
        from sqlalchemy import select

        return list(
            session.scalars(
                select(EvaluationConfiguration).where(
                    EvaluationConfiguration.tenant_id == tenant_id,
                    EvaluationConfiguration.customized_workflow_id == customized_workflow_id,
                )
            ).all()
        )

    # ---- Evaluation Run Management ----

    @classmethod
    def start_evaluation_run(
        cls,
        session: Session,
        tenant_id: str,
        target_type: str,
        target_id: str,
        account_id: str,
        dataset_file_content: bytes,
        dataset_filename: str,
        run_request: EvaluationRunRequest,
    ) -> EvaluationRun:
        """Validate dataset, create run record, dispatch Celery task.

        Saves the provided parameters as the latest EvaluationConfiguration
        before creating the run.
        """
        # Check framework is configured
        evaluation_instance = EvaluationManager.get_evaluation_instance()
        if evaluation_instance is None:
            raise EvaluationFrameworkNotConfiguredError()

        # Save as latest EvaluationConfiguration
        config = cls.save_evaluation_config(
            session=session,
            tenant_id=tenant_id,
            target_type=target_type,
            target_id=target_id,
            account_id=account_id,
            data=run_request,
        )

        # Check concurrent run limit
        active_runs = (
            session.query(EvaluationRun)
            .filter_by(tenant_id=tenant_id)
            .filter(EvaluationRun.status.in_([EvaluationRunStatus.PENDING, EvaluationRunStatus.RUNNING]))
            .count()
        )
        max_concurrent = dify_config.EVALUATION_MAX_CONCURRENT_RUNS
        if active_runs >= max_concurrent:
            raise EvaluationMaxConcurrentRunsError(f"Maximum concurrent runs ({max_concurrent}) reached.")

        # Parse dataset
        items = cls._parse_dataset(dataset_file_content, dataset_filename)
        max_rows = dify_config.EVALUATION_MAX_DATASET_ROWS
        if len(items) > max_rows:
            raise EvaluationDatasetInvalidError(f"Dataset has {len(items)} rows, max is {max_rows}.")

        # Create evaluation run
        evaluation_run = EvaluationRun(
            tenant_id=tenant_id,
            target_type=target_type,
            target_id=target_id,
            evaluation_config_id=config.id,
            status=EvaluationRunStatus.PENDING,
            dataset_file_id=run_request.file_id,
            total_items=len(items),
            created_by=account_id,
        )
        session.add(evaluation_run)
        session.commit()
        session.refresh(evaluation_run)

        # Build Celery task data
        run_data = EvaluationRunData(
            evaluation_run_id=evaluation_run.id,
            tenant_id=tenant_id,
            target_type=target_type,
            target_id=target_id,
            evaluation_model_provider=run_request.evaluation_model_provider,
            evaluation_model=run_request.evaluation_model,
            default_metrics=run_request.default_metrics,
            customized_metrics=run_request.customized_metrics,
            judgment_config=run_request.judgment_config,
            input_list=items,
        )

        # Dispatch Celery task
        from tasks.evaluation_task import run_evaluation

        task = run_evaluation.delay(run_data.model_dump())
        evaluation_run.celery_task_id = task.id
        session.commit()

        return evaluation_run

    @classmethod
    def start_stub_evaluation_run(
        cls,
        session: Session,
        tenant_id: str,
        target_type: str,
        target_id: str,
        account_id: str,
        dataset_file_content: bytes,
        dataset_filename: str,
        run_request: EvaluationRunRequest,
    ) -> EvaluationRun:
        """Persist a completed synthetic run for frontend integration testing.

        This temporary path keeps the existing read flows (`logs`, `run detail`,
        and result-file download) working for app evaluations while the real
        execution logic is moved to `/evaluation/run1` for backend iteration.
        """
        from tasks.evaluation_task import (
            _compute_metrics_summary,
            _generate_result_xlsx,
            _persist_results,
            _store_result_file,
        )

        config = cls.save_evaluation_config(
            session=session,
            tenant_id=tenant_id,
            target_type=target_type,
            target_id=target_id,
            account_id=account_id,
            data=run_request,
        )

        items = cls._parse_dataset(dataset_file_content, dataset_filename)
        max_rows = dify_config.EVALUATION_MAX_DATASET_ROWS
        if len(items) > max_rows:
            raise EvaluationDatasetInvalidError(f"Dataset has {len(items)} rows, max is {max_rows}.")

        now = naive_utc_now()
        results = cls._build_stub_results(input_list=items, run_request=run_request)
        metrics_summary = _compute_metrics_summary(results, run_request.judgment_config)

        evaluation_run = EvaluationRun(
            tenant_id=tenant_id,
            target_type=target_type,
            target_id=target_id,
            evaluation_config_id=config.id,
            status=EvaluationRunStatus.COMPLETED,
            dataset_file_id=run_request.file_id,
            total_items=len(items),
            completed_items=len(items),
            failed_items=0,
            metrics_summary=json.dumps(metrics_summary),
            created_by=account_id,
            started_at=now,
            completed_at=now,
        )
        session.add(evaluation_run)
        session.commit()
        session.refresh(evaluation_run)

        _persist_results(session, evaluation_run.id, results, items)

        result_xlsx = _generate_result_xlsx(items, results)
        result_file_id = _store_result_file(
            tenant_id,
            evaluation_run.id,
            result_xlsx,
            session,
            created_by=account_id,
        )
        if result_file_id:
            evaluation_run.result_file_id = result_file_id
            session.commit()
            session.refresh(evaluation_run)

        return evaluation_run

    @classmethod
    def get_evaluation_runs(
        cls,
        session: Session,
        tenant_id: str,
        target_type: str,
        target_id: str,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[EvaluationRun], int]:
        """Query evaluation run history with pagination."""
        query = (
            session.query(EvaluationRun)
            .filter_by(tenant_id=tenant_id, target_type=target_type, target_id=target_id)
            .order_by(EvaluationRun.created_at.desc())
        )
        total = query.count()
        runs = query.offset((page - 1) * page_size).limit(page_size).all()
        return runs, total

    @classmethod
    def get_evaluation_run_detail(
        cls,
        session: Session,
        tenant_id: str,
        run_id: str,
    ) -> EvaluationRun:
        run = session.query(EvaluationRun).filter_by(id=run_id, tenant_id=tenant_id).first()
        if not run:
            raise EvaluationNotFoundError("Evaluation run not found.")
        return run

    @classmethod
    def get_evaluation_run_items(
        cls,
        session: Session,
        run_id: str,
        page: int = 1,
        page_size: int = 50,
    ) -> tuple[list[EvaluationRunItem], int]:
        """Query evaluation run items with pagination."""
        query = (
            session.query(EvaluationRunItem)
            .filter_by(evaluation_run_id=run_id)
            .order_by(EvaluationRunItem.item_index.asc())
        )
        total = query.count()
        items = query.offset((page - 1) * page_size).limit(page_size).all()
        return items, total

    @classmethod
    def cancel_evaluation_run(
        cls,
        session: Session,
        tenant_id: str,
        run_id: str,
    ) -> EvaluationRun:
        run = cls.get_evaluation_run_detail(session, tenant_id, run_id)
        if run.status not in (EvaluationRunStatus.PENDING, EvaluationRunStatus.RUNNING):
            raise ValueError(f"Cannot cancel evaluation run in status: {run.status}")

        run.status = EvaluationRunStatus.CANCELLED

        # Revoke Celery task if running
        if run.celery_task_id:
            try:
                from celery import current_app as celery_app

                celery_app.control.revoke(run.celery_task_id, terminate=True)
            except Exception:
                logger.exception("Failed to revoke Celery task %s", run.celery_task_id)

        session.commit()
        return run

    @classmethod
    def get_supported_metrics(cls, category: EvaluationCategory) -> list[str]:
        return EvaluationManager.get_supported_metrics(category)

    @classmethod
    def get_available_metrics(cls) -> list[str]:
        """Return the centrally-defined list of evaluation metrics."""
        return [m.value for m in EvaluationMetricName]

    @classmethod
    def filter_console_default_metrics(
        cls,
        default_metrics: list[DefaultMetric | Mapping[str, Any]],
    ) -> list[DefaultMetric]:
        """Drop agent-only metrics/nodes from console-facing evaluation configs."""
        filtered: list[DefaultMetric] = []
        for raw_metric in default_metrics:
            metric = raw_metric if isinstance(raw_metric, DefaultMetric) else DefaultMetric.model_validate(raw_metric)
            if metric.metric in cls.CONSOLE_DISABLED_METRICS:
                continue

            visible_nodes = [node for node in metric.node_info_list if node.type not in cls.CONSOLE_DISABLED_CATEGORIES]
            if not visible_nodes:
                continue

            filtered.append(metric.model_copy(update={"node_info_list": visible_nodes}))
        return filtered

    @classmethod
    def filter_console_judgment_config(
        cls,
        judgment_config: JudgmentConfig | Mapping[str, Any] | None,
    ) -> JudgmentConfig | None:
        """Strip judgment conditions that reference metrics hidden from console evaluation."""
        if judgment_config is None:
            return None

        config = (
            judgment_config
            if isinstance(judgment_config, JudgmentConfig)
            else JudgmentConfig.model_validate(judgment_config)
        )
        visible_conditions = [
            condition
            for condition in config.conditions
            if len(condition.variable_selector) < 2 or condition.variable_selector[1] not in cls.CONSOLE_DISABLED_METRICS
        ]
        if not visible_conditions:
            return None
        return config.model_copy(update={"conditions": visible_conditions})

    @classmethod
    def serialize_console_default_metrics(
        cls,
        default_metrics: list[Mapping[str, Any]],
    ) -> list[dict[str, Any]]:
        """Return sanitized default metrics payload for console responses."""
        return [metric.model_dump() for metric in cls.filter_console_default_metrics(default_metrics)]

    @classmethod
    def serialize_console_judgment_config(
        cls,
        judgment_config: Mapping[str, Any] | None,
    ) -> dict[str, Any] | None:
        """Return sanitized judgment config payload for console responses."""
        filtered = cls.filter_console_judgment_config(judgment_config)
        return filtered.model_dump() if filtered else None

    @classmethod
    def get_dataset_column_names(
        cls,
        target: Union[App, CustomizedSnippet],
        target_type: str,
        data: EvaluationConfigData,
    ) -> list[str]:
        """Build dataset column names from target inputs and the selected evaluation config."""
        input_columns = cls._get_target_input_column_names(target, target_type)
        expected_output_columns = cls._get_expected_output_column_names(data.default_metrics)
        return ["index", *input_columns, *expected_output_columns]

    @classmethod
    def _get_target_input_column_names(
        cls,
        target: Union[App, CustomizedSnippet],
        target_type: str,
    ) -> list[str]:
        """Resolve user-input variables for the target in workflow order."""
        if target_type == EvaluationTargetType.APPS.value and isinstance(target, App):
            input_fields = cls._get_app_input_fields(target)
        elif target_type == EvaluationTargetType.SNIPPETS.value and isinstance(target, CustomizedSnippet):
            input_fields = cls._get_snippet_input_fields(target)
        else:
            raise ValueError(f"Unsupported target type: {target_type}")

        columns: list[str] = []
        seen: set[str] = set()
        for field in input_fields:
            column_name = str(field.get("variable") or field.get("label") or "").strip()
            if not column_name or column_name in seen:
                continue
            seen.add(column_name)
            columns.append(column_name)
        return columns

    @classmethod
    def _get_expected_output_column_names(
        cls,
        default_metrics: list[DefaultMetric | Mapping[str, Any]],
    ) -> list[str]:
        """Build one expected_output column per visible node that needs a reference answer."""
        columns: list[str] = []
        seen: set[str] = set()
        for metric in cls.filter_console_default_metrics(default_metrics):
            if metric.metric not in cls.METRICS_REQUIRING_EXPECTED_OUTPUT:
                continue
            for node_info in metric.node_info_list:
                node_title = node_info.title or node_info.node_id
                column_name = f"{node_title} : expected_output"
                if column_name in seen:
                    continue
                seen.add(column_name)
                columns.append(column_name)
        return columns

    @classmethod
    def _nodes_for_metrics_from_workflow(
        cls,
        workflow: Workflow,
        metrics: list[str],
    ) -> dict[str, list[dict[str, str]]]:
        node_type_to_nodes: dict[str, list[dict[str, str]]] = {}
        for node_id, node_data in workflow.walk_nodes():
            ntype = node_data.get("type", "")
            if ntype in cls.CONSOLE_DISABLED_CATEGORIES:
                continue
            node_type_to_nodes.setdefault(ntype, []).append(
                NodeInfo(node_id=node_id, type=ntype, title=node_data.get("title", "")).model_dump()
            )

        result: dict[str, list[dict[str, str]]] = {}
        for metric in metrics:
            required_node_type = METRIC_NODE_TYPE_MAPPING.get(metric)
            if required_node_type is None:
                result[metric] = []
                continue
            result[metric] = node_type_to_nodes.get(required_node_type, [])
        return result

    @classmethod
    def _union_supported_metric_names(cls) -> list[str]:
        """Metric names the current evaluation framework supports for any :class:`EvaluationCategory`."""
        ordered: list[str] = []
        seen: set[str] = set()
        for category in EvaluationCategory:
            for name in cls.get_supported_metrics(category):
                if name not in seen:
                    seen.add(name)
                    ordered.append(name)
        return ordered

    @classmethod
    def get_default_metrics_with_nodes_for_published_target(
        cls,
        target: Union[App, CustomizedSnippet],
        target_type: str,
    ) -> list[DefaultMetric]:
        """List default metrics and matching nodes using only the *published* workflow graph.

        Metrics are those supported by the configured evaluation framework and present in
        :data:`METRIC_NODE_TYPE_MAPPING`. Node lists are derived from the published workflow only
        (no draft fallback).
        """
        workflow = cls._resolve_published_workflow(target, target_type)
        if not workflow:
            return []

        supported = cls._union_supported_metric_names()
        metric_names = sorted(m for m in supported if m in METRIC_NODE_TYPE_MAPPING)
        if not metric_names:
            return []

        nodes_by_metric = cls._nodes_for_metrics_from_workflow(workflow, metric_names)
        return [
            DefaultMetric(
                metric=m,
                value_type=METRIC_VALUE_TYPE_MAPPING.get(m, "number"),
                node_info_list=[NodeInfo.model_validate(n) for n in nodes_by_metric.get(m, [])],
            )
            for m in metric_names
        ]

    @classmethod
    def get_nodes_for_metrics(
        cls,
        target: Union[App, CustomizedSnippet],
        target_type: str,
        metrics: list[str] | None = None,
    ) -> dict[str, list[dict[str, str]]]:
        """Return node info grouped by metric (or all nodes when *metrics* is empty).

        :param target: App or CustomizedSnippet instance.
        :param target_type: ``"apps"`` or ``"snippets"``.
        :param metrics: Optional list of metric names to filter by.
            When *None* or empty, returns ``{"all": [<every node>]}``.
        :returns: ``{metric_name: [NodeInfo dict, ...]}`` or
            ``{"all": [NodeInfo dict, ...]}``.
        """
        workflow = cls._resolve_workflow(target, target_type)
        if not workflow:
            return {"all": []} if not metrics else {m: [] for m in metrics}

        if not metrics:
            all_nodes = [
                NodeInfo(node_id=node_id, type=node_data.get("type", ""), title=node_data.get("title", "")).model_dump()
                for node_id, node_data in workflow.walk_nodes()
                if node_data.get("type", "") not in cls.CONSOLE_DISABLED_CATEGORIES
            ]
            return {"all": all_nodes}

        return cls._nodes_for_metrics_from_workflow(workflow, metrics)

    @classmethod
    def _resolve_published_workflow(
        cls,
        target: Union[App, CustomizedSnippet],
        target_type: str,
    ) -> Workflow | None:
        """Resolve only the published workflow for the target (no draft fallback)."""
        if target_type == EvaluationTargetType.SNIPPETS.value and isinstance(target, CustomizedSnippet):
            return SnippetService().get_published_workflow(snippet=target)
        if target_type == EvaluationTargetType.APPS.value and isinstance(target, App):
            return WorkflowService().get_published_workflow(app_model=target)
        return None

    @classmethod
    def _resolve_workflow(
        cls,
        target: Union[App, CustomizedSnippet],
        target_type: str,
    ) -> Workflow | None:
        """Resolve the *published* (preferred) or *draft* workflow for the target."""
        if target_type == EvaluationTargetType.SNIPPETS.value and isinstance(target, CustomizedSnippet):
            snippet_service = SnippetService()
            workflow = snippet_service.get_published_workflow(snippet=target)
            if not workflow:
                workflow = snippet_service.get_draft_workflow(snippet=target)
            return workflow
        elif target_type == EvaluationTargetType.APPS.value and isinstance(target, App):
            workflow_service = WorkflowService()
            workflow = workflow_service.get_published_workflow(app_model=target)
            if not workflow:
                workflow = workflow_service.get_draft_workflow(app_model=target)
            return workflow
        return None

    # ---- Category Resolution ----

    @classmethod
    def _resolve_evaluation_category(cls, default_metrics: list[DefaultMetric]) -> EvaluationCategory:
        """Derive evaluation category from default_metrics node_info types.

        Uses the type of the first node_info found in default_metrics.
        Falls back to LLM if no metrics are provided.
        """
        for metric in default_metrics:
            for node_info in metric.node_info_list:
                try:
                    return EvaluationCategory(node_info.type)
                except ValueError:
                    continue
        return EvaluationCategory.LLM

    @classmethod
    def execute_targets(
        cls,
        tenant_id: str,
        target_type: str,
        target_id: str,
        input_list: list[EvaluationDatasetInput],
        max_workers: int = 5,
    ) -> tuple[list[dict[str, NodeRunResult]], list[str | None]]:
        """Execute the evaluation target for every test-data item in parallel.

        :param tenant_id: Workspace / tenant ID.
        :param target_type: ``"apps"`` or ``"snippets"``.
        :param target_id: ID of the App or CustomizedSnippet.
        :param input_list: All test-data items parsed from the dataset.
        :param max_workers: Maximum number of parallel worker threads.
        :return: Tuple of (node_results, workflow_run_ids).
            node_results: ordered list of ``{node_id: NodeRunResult}`` mappings;
            the *i*-th element corresponds to ``input_list[i]``.
            workflow_run_ids: ordered list of workflow_run_id strings (or None)
            for each input item.
        """
        from concurrent.futures import ThreadPoolExecutor

        from flask import Flask, current_app

        flask_app: Flask = current_app._get_current_object()  # type: ignore

        def _worker(item: EvaluationDatasetInput) -> tuple[dict[str, NodeRunResult], str | None]:
            with flask_app.app_context():
                from models.engine import db

                with Session(db.engine, expire_on_commit=False) as thread_session:
                    try:
                        response = cls._run_single_target(
                            session=thread_session,
                            target_type=target_type,
                            target_id=target_id,
                            item=item,
                        )

                        workflow_run_id = cls._extract_workflow_run_id(response)
                        if not workflow_run_id:
                            logger.warning(
                                "No workflow_run_id for item %d (target=%s)",
                                item.index,
                                target_id,
                            )
                            return {}, None

                        node_results = cls._query_node_run_results(
                            session=thread_session,
                            tenant_id=tenant_id,
                            app_id=target_id,
                            workflow_run_id=workflow_run_id,
                        )
                        return node_results, workflow_run_id
                    except Exception:
                        logger.exception(
                            "Target execution failed for item %d (target=%s)",
                            item.index,
                            target_id,
                        )
                        return {}, None

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(_worker, item) for item in input_list]
            ordered_results: list[dict[str, NodeRunResult]] = []
            ordered_workflow_run_ids: list[str | None] = []
            for future in futures:
                try:
                    node_result, wf_run_id = future.result()
                    ordered_results.append(node_result)
                    ordered_workflow_run_ids.append(wf_run_id)
                except Exception:
                    logger.exception("Unexpected error collecting target execution result")
                    ordered_results.append({})
                    ordered_workflow_run_ids.append(None)

        return ordered_results, ordered_workflow_run_ids

    @classmethod
    def _run_single_target(
        cls,
        session: Session,
        target_type: str,
        target_id: str,
        item: EvaluationDatasetInput,
    ) -> Mapping[str, object]:
        """Execute a single evaluation target with one test-data item.

        Dispatches to the appropriate execution service based on
        ``target_type``:

        * ``"snippets"`` → :meth:`SnippetGenerateService.run_published`
        * ``"apps"`` → :meth:`WorkflowAppGenerator().generate` (blocking mode)

        :returns: The blocking response mapping from the workflow engine.
        :raises ValueError: If the target is not found or not published.
        """
        from core.app.apps.workflow.app_generator import WorkflowAppGenerator
        from core.app.entities.app_invoke_entities import InvokeFrom
        from core.evaluation.runners import get_service_account_for_app, get_service_account_for_snippet

        if target_type == EvaluationTargetType.SNIPPETS.value:
            from services.snippet_generate_service import SnippetGenerateService

            snippet = session.query(CustomizedSnippet).filter_by(id=target_id).first()
            if not snippet:
                raise ValueError(f"Snippet {target_id} not found")

            service_account = get_service_account_for_snippet(session, target_id)

            return SnippetGenerateService.run_published(
                snippet=snippet,
                user=service_account,
                args={"inputs": item.inputs},
                invoke_from=InvokeFrom.SERVICE_API,
            )
        else:
            # target_type == "apps"
            app = session.query(App).filter_by(id=target_id).first()
            if not app:
                raise ValueError(f"App {target_id} not found")

            service_account = get_service_account_for_app(session, target_id)

            workflow_service = WorkflowService()
            workflow = workflow_service.get_published_workflow(app_model=app)
            if not workflow:
                raise ValueError(f"No published workflow for app {target_id}")

            response: Mapping[str, object] = WorkflowAppGenerator().generate(
                app_model=app,
                workflow=workflow,
                user=service_account,
                args={"inputs": item.inputs},
                invoke_from=InvokeFrom.SERVICE_API,
                streaming=False,
                call_depth=0,
            )
            return response

    @staticmethod
    def _extract_workflow_run_id(response: Mapping[str, object]) -> str | None:
        """Extract ``workflow_run_id`` from a blocking workflow response."""
        wf_run_id = response.get("workflow_run_id")
        if wf_run_id:
            return str(wf_run_id)
        data = response.get("data")
        if isinstance(data, Mapping) and data.get("id"):
            return str(data["id"])
        return None

    @staticmethod
    def _query_node_run_results(
        session: Session,
        tenant_id: str,
        app_id: str,
        workflow_run_id: str,
    ) -> dict[str, NodeRunResult]:
        """Query all node execution records for a workflow run."""
        from sqlalchemy import asc, select

        from graphon.enums import WorkflowNodeExecutionStatus
        from models.workflow import WorkflowNodeExecutionModel

        stmt = (
            WorkflowNodeExecutionModel.preload_offload_data(select(WorkflowNodeExecutionModel))
            .where(
                WorkflowNodeExecutionModel.tenant_id == tenant_id,
                WorkflowNodeExecutionModel.app_id == app_id,
                WorkflowNodeExecutionModel.workflow_run_id == workflow_run_id,
            )
            .order_by(asc(WorkflowNodeExecutionModel.created_at))
        )

        node_models: list[WorkflowNodeExecutionModel] = list(session.execute(stmt).scalars().all())

        result: dict[str, NodeRunResult] = {}
        for node in node_models:
            # Convert string-keyed metadata to WorkflowNodeExecutionMetadataKey-keyed
            raw_metadata = node.execution_metadata_dict
            typed_metadata: dict[WorkflowNodeExecutionMetadataKey, object] = {}
            for key, val in raw_metadata.items():
                try:
                    typed_metadata[WorkflowNodeExecutionMetadataKey(key)] = val
                except ValueError:
                    pass  # skip unknown metadata keys

            result[node.node_id] = NodeRunResult(
                status=WorkflowNodeExecutionStatus(node.status),
                inputs=node.inputs_dict or {},
                process_data=node.process_data_dict or {},
                outputs=node.outputs_dict or {},
                metadata=typed_metadata,
                error=node.error or "",
            )
        return result

    # ---- Dataset Parsing ----

    @classmethod
    def _parse_dataset(cls, file_content: bytes, filename: str) -> list[EvaluationDatasetInput]:
        """Parse evaluation dataset from CSV or XLSX content.

        Supported schemas:

        - ``index,<input...>,expected_output`` for the legacy single-reference flow
        - ``index,<input...>,<node title> : expected_output`` for node-specific references

        Column parsing treats the *last* ``:`` as the separator so node titles
        can themselves contain ``:`` characters.
        """
        filename_lower = filename.lower()
        if filename_lower.endswith(".csv"):
            return cls._parse_csv_dataset(file_content)
        return cls._parse_xlsx_dataset(file_content)

    @classmethod
    def _parse_xlsx_dataset(cls, xlsx_content: bytes) -> list[EvaluationDatasetInput]:
        """Parse evaluation dataset from XLSX bytes."""
        wb = load_workbook(io.BytesIO(xlsx_content), read_only=True)
        ws = wb.active
        if ws is None:
            raise EvaluationDatasetInvalidError("XLSX file has no active worksheet.")

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise EvaluationDatasetInvalidError("Dataset must have at least a header row and one data row.")

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        if not headers or headers[0].lower() != "index":
            raise EvaluationDatasetInvalidError("First column header must be 'index'.")

        items = cls._rows_to_dataset_items(rows)
        wb.close()
        return items

    @classmethod
    def _parse_csv_dataset(cls, csv_content: bytes) -> list[EvaluationDatasetInput]:
        """Parse evaluation dataset from UTF-8 CSV bytes.

        CSV follows the same schema as XLSX:
        the first column must be `index`, remaining columns become inputs,
        and expected-output columns can be either the legacy
        ``expected_output`` or node-scoped ``<node title> : expected_output``.
        """
        try:
            decoded = csv_content.decode("utf-8-sig")
        except UnicodeDecodeError as e:
            raise EvaluationDatasetInvalidError("CSV file must be UTF-8 encoded.") from e

        reader = csv.reader(io.StringIO(decoded))
        rows = list(reader)
        if len(rows) < 2:
            raise EvaluationDatasetInvalidError("Dataset must have at least a header row and one data row.")

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        if not headers or headers[0].lower() != "index":
            raise EvaluationDatasetInvalidError("First column header must be 'index'.")

        return cls._rows_to_dataset_items(rows)

    @classmethod
    def _rows_to_dataset_items(cls, rows: list[list[Any] | tuple[Any, ...]]) -> list[EvaluationDatasetInput]:
        """Normalize spreadsheet rows into dataset items with index validation."""
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        input_headers = headers[1:]

        items: list[EvaluationDatasetInput] = []
        seen_indices: set[int] = set()
        for row_idx, row in enumerate(rows[1:], start=1):
            values = list(row)
            if all(v is None or str(v).strip() == "" for v in values):
                continue

            index_val = values[0] if values else row_idx
            try:
                index = int(str(index_val))
            except (TypeError, ValueError):
                index = row_idx

            if index in seen_indices:
                raise EvaluationDatasetInvalidError(f"Dataset index '{index}' is duplicated.")
            seen_indices.add(index)

            inputs: dict[str, Any] = {}
            expected_output: str | None = None
            expected_outputs: dict[str, str] = {}
            for col_idx, header in enumerate(input_headers):
                val = values[col_idx + 1] if col_idx + 1 < len(values) else None
                string_value = str(val) if val is not None else ""
                expected_output_target = cls._parse_expected_output_header(header)
                if expected_output_target is None:
                    inputs[header] = string_value
                    continue

                if expected_output_target == "":
                    expected_output = string_value
                else:
                    expected_outputs[expected_output_target] = string_value

            items.append(
                EvaluationDatasetInput(
                    index=index,
                    inputs=inputs,
                    expected_output=expected_output,
                    expected_outputs=expected_outputs,
                )
            )

        return items

    @staticmethod
    def _parse_expected_output_header(header: str) -> str | None:
        """Return the node title for an expected-output column, if any.

        ``expected_output`` keeps the legacy single-reference behaviour and
        returns an empty string marker. Node-specific columns split on the last
        ``:`` so titles such as ``RAG: final judge`` remain valid.
        """
        normalized_header = header.strip()
        if normalized_header == "expected_output":
            return ""

        if ":" not in normalized_header:
            return None

        node_title, suffix = normalized_header.rsplit(":", 1)
        if suffix.strip() != "expected_output":
            return None

        normalized_title = node_title.strip()
        return normalized_title or None

    @classmethod
    def _build_stub_results(
        cls,
        input_list: list[EvaluationDatasetInput],
        run_request: EvaluationRunRequest,
    ) -> list[EvaluationItemResult]:
        """Create deterministic synthetic results that match the real read models."""
        results: list[EvaluationItemResult] = []

        for item_position, item in enumerate(input_list):
            metrics: list[EvaluationMetric] = []

            for metric_position, default_metric in enumerate(run_request.default_metrics):
                metric_value_type = default_metric.value_type or METRIC_VALUE_TYPE_MAPPING.get(
                    default_metric.metric, ""
                )
                for node_position, node_info in enumerate(default_metric.node_info_list):
                    metrics.append(
                        EvaluationMetric(
                            name=default_metric.metric,
                            value=cls._build_stub_metric_value(
                                item_index=item.index,
                                metric_position=metric_position,
                                node_position=node_position,
                                value_type=metric_value_type,
                                metric_name=default_metric.metric,
                            ),
                            details={
                                "stubbed": True,
                                "source": "console-evaluation-run",
                                "value_type": metric_value_type,
                            },
                            node_info=node_info,
                        )
                    )

            if run_request.customized_metrics:
                for output_position, output_field in enumerate(run_request.customized_metrics.output_fields):
                    metrics.append(
                        EvaluationMetric(
                            name=output_field.variable,
                            value=cls._build_stub_metric_value(
                                item_index=item.index,
                                metric_position=len(metrics),
                                node_position=output_position,
                                value_type=output_field.value_type,
                                metric_name=output_field.variable,
                            ),
                            details={
                                "stubbed": True,
                                "source": "console-evaluation-run",
                                "value_type": output_field.value_type,
                                "customized": True,
                            },
                            node_info=NodeInfo(
                                node_id=run_request.customized_metrics.evaluation_workflow_id,
                                type="customized",
                                title="customized",
                            ),
                        )
                    )

            judgment = cls._evaluate_stub_judgment(metrics, run_request.judgment_config)
            if isinstance(item, EvaluationDatasetInput):
                actual_output = item.get_expected_output_for_node(None)
            else:
                actual_output = getattr(item, "expected_output", None)
            actual_output = actual_output or cls._build_stub_output(item_position, item.inputs)

            results.append(
                EvaluationItemResult(
                    index=item.index,
                    actual_output=actual_output,
                    metrics=metrics,
                    metadata={
                        "stubbed": True,
                        "source": "console-evaluation-run",
                        "row": item_position + 1,
                    },
                    judgment=judgment,
                )
            )

        return results

    @staticmethod
    def _build_stub_output(item_position: int, inputs: dict[str, Any]) -> str:
        """Build a readable synthetic output for one dataset row."""
        first_key = next(iter(inputs.keys()), "input")
        first_value = inputs.get(first_key, "")
        return f"Stub output #{item_position + 1}: processed {first_key}={first_value}"

    @staticmethod
    def _build_stub_metric_value(
        item_index: int,
        metric_position: int,
        node_position: int,
        value_type: str,
        metric_name: str,
    ) -> Any:
        """Return a deterministic placeholder metric value by declared type."""
        base_seed = item_index + metric_position + node_position
        normalized_type = value_type.lower()

        if normalized_type == "number" or not normalized_type:
            return round(0.72 + (base_seed % 18) / 100, 3)
        if normalized_type == "boolean":
            return base_seed % 2 == 0

        return f"stub-{metric_name}-{item_index}"

    @staticmethod
    def _evaluate_stub_judgment(
        metrics: list[EvaluationMetric],
        judgment_config: JudgmentConfig | None,
    ) -> JudgmentResult:
        """Apply the same judgment processor used by real evaluations."""
        if not judgment_config or not judgment_config.conditions:
            return JudgmentResult()

        metric_values: dict[tuple[str, str], object] = {
            (metric.node_info.node_id, metric.name): metric.value for metric in metrics if metric.node_info
        }
        return JudgmentProcessor.evaluate(metric_values, judgment_config)

    @classmethod
    def execute_retrieval_test_targets(
        cls,
        dataset_id: str,
        account_id: str,
        input_list: list[EvaluationDatasetInput],
        max_workers: int = 5,
    ) -> list[NodeRunResult]:
        """Run hit testing against a knowledge base for every input item in parallel.

        Each item must supply a ``query`` key in its ``inputs`` dict.  The
        retrieved segments are normalised into the same ``NodeRunResult`` format
        that :class:`RetrievalEvaluationRunner` expects:

        .. code-block:: python

            NodeRunResult(
                inputs={"query": "..."},
                outputs={"result": [{"content": "...", "score": ...}, ...]},
            )

        :returns: Ordered list of ``NodeRunResult`` — one per input item.
            If retrieval fails for an item the result has an empty ``result``
            list so the runner can still persist a (metric-less) row.
        """
        from concurrent.futures import ThreadPoolExecutor

        from flask import current_app

        flask_app = current_app._get_current_object()  # type: ignore

        def _worker(item: EvaluationDatasetInput) -> NodeRunResult:
            with flask_app.app_context():
                from extensions.ext_database import db as flask_db
                from models.account import Account
                from models.dataset import Dataset
                from services.hit_testing_service import HitTestingService

                dataset = flask_db.session.query(Dataset).filter_by(id=dataset_id).first()
                if not dataset:
                    raise ValueError(f"Dataset {dataset_id} not found")

                account = flask_db.session.query(Account).filter_by(id=account_id).first()
                if not account:
                    raise ValueError(f"Account {account_id} not found")

                query = str(item.inputs.get("query", ""))
                response = HitTestingService.retrieve(
                    dataset=dataset,
                    query=query,
                    account=account,
                    retrieval_model=None,  # Use dataset's configured retrieval model
                    external_retrieval_model={},
                    limit=10,
                )

                records = response.get("records", [])
                result_list = [
                    {
                        "content": r.get("segment", {}).get("content", "") or r.get("content", ""),
                        "score": r.get("score"),
                    }
                    for r in records
                    if r.get("segment", {}).get("content") or r.get("content")
                ]

                return NodeRunResult(
                    inputs={"query": query},
                    outputs={"result": result_list},
                )

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(_worker, item) for item in input_list]
            results: list[NodeRunResult] = []
            for item, future in zip(input_list, futures):
                try:
                    results.append(future.result())
                except Exception:
                    logger.exception("Retrieval test failed for item %d (dataset=%s)", item.index, dataset_id)
                    results.append(NodeRunResult(inputs={}, outputs={"result": []}))

        return results
