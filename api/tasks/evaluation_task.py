import io
import json
import logging
from typing import Any

from celery import shared_task
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from configs import dify_config
from core.evaluation.base_evaluation_instance import BaseEvaluationInstance
from core.evaluation.entities.evaluation_entity import (
    EvaluationCategory,
    EvaluationDatasetInput,
    EvaluationItemResult,
    EvaluationRunData,
    NodeInfo,
)
from core.evaluation.entities.judgment_entity import JudgmentConfig
from core.evaluation.evaluation_manager import EvaluationManager
from core.evaluation.judgment.processor import JudgmentProcessor
from core.evaluation.runners.agent_evaluation_runner import AgentEvaluationRunner
from core.evaluation.runners.base_evaluation_runner import BaseEvaluationRunner
from core.evaluation.runners.llm_evaluation_runner import LLMEvaluationRunner
from core.evaluation.runners.retrieval_evaluation_runner import RetrievalEvaluationRunner
from core.evaluation.runners.snippet_evaluation_runner import SnippetEvaluationRunner
from core.evaluation.runners.workflow_evaluation_runner import WorkflowEvaluationRunner
from extensions.ext_database import db
from graphon.node_events import NodeRunResult
from libs.datetime_utils import naive_utc_now
from models.enums import CreatorUserRole
from models.evaluation import EvaluationRun, EvaluationRunItem, EvaluationRunStatus
from models.model import UploadFile
from services.evaluation_service import EvaluationService

logger = logging.getLogger(__name__)


@shared_task(queue="evaluation")
def run_evaluation(run_data_dict: dict[str, Any]) -> None:
    """Celery task for running evaluations asynchronously.

    Workflow:
    1. Deserialize EvaluationRunData
    2. Execute target and collect node results
    3. Evaluate metrics via runners (one per metric-node pair)
    4. Merge results per test-data row (1 item = 1 EvaluationRunItem)
    5. Apply judgment conditions
    6. Persist results + generate result XLSX
    7. Update EvaluationRun status to COMPLETED
    """
    run_data = EvaluationRunData.model_validate(run_data_dict)

    with db.engine.connect() as connection:
        from sqlalchemy.orm import Session

        session = Session(bind=connection)

        try:
            _execute_evaluation(session, run_data)
        except Exception as e:
            logger.exception("Evaluation run %s failed", run_data.evaluation_run_id)
            _mark_run_failed(session, run_data.evaluation_run_id, str(e))
        finally:
            session.close()


def _execute_evaluation(session: Any, run_data: EvaluationRunData) -> None:
    """Core evaluation execution logic."""
    evaluation_run = session.query(EvaluationRun).filter_by(id=run_data.evaluation_run_id).first()
    if not evaluation_run:
        logger.error("EvaluationRun %s not found", run_data.evaluation_run_id)
        return

    if evaluation_run.status == EvaluationRunStatus.CANCELLED:
        logger.info("EvaluationRun %s was cancelled", run_data.evaluation_run_id)
        return

    evaluation_instance = EvaluationManager.get_evaluation_instance()
    if evaluation_instance is None:
        raise ValueError("Evaluation framework not configured")

    # Mark as running
    evaluation_run.status = EvaluationRunStatus.RUNNING
    evaluation_run.started_at = naive_utc_now()
    session.commit()

    if run_data.target_type == "dataset":
        results: list[EvaluationItemResult] = _execute_retrieval_test(
            session=session,
            evaluation_run=evaluation_run,
            run_data=run_data,
            evaluation_instance=evaluation_instance,
        )
    else:
        evaluation_service = EvaluationService()
        node_run_result_mapping_list, workflow_run_ids = evaluation_service.execute_targets(
            tenant_id=run_data.tenant_id,
            target_type=run_data.target_type,
            target_id=run_data.target_id,
            input_list=run_data.input_list,
        )

        workflow_run_id_map = {
            item.index: wf_run_id
            for item, wf_run_id in zip(run_data.input_list, workflow_run_ids)
            if wf_run_id
        }

        results = _execute_evaluation_runner(
            session=session,
            run_data=run_data,
            evaluation_instance=evaluation_instance,
            node_run_result_mapping_list=node_run_result_mapping_list,
            workflow_run_id_map=workflow_run_id_map,
        )

    # Compute summary metrics
    metrics_summary = _compute_metrics_summary(results, run_data.judgment_config)

    # Generate result XLSX
    result_xlsx = _generate_result_xlsx(run_data.input_list, results)

    # Store result file
    result_file_id = _store_result_file(
        run_data.tenant_id,
        run_data.evaluation_run_id,
        result_xlsx,
        session,
        created_by=evaluation_run.created_by,
    )

    # Update run to completed
    evaluation_run = session.query(EvaluationRun).filter_by(id=run_data.evaluation_run_id).first()
    if evaluation_run:
        evaluation_run.status = EvaluationRunStatus.COMPLETED
        evaluation_run.completed_at = naive_utc_now()
        evaluation_run.metrics_summary = json.dumps(metrics_summary)
        if result_file_id:
            evaluation_run.result_file_id = result_file_id
        session.commit()

    logger.info("Evaluation run %s completed successfully", run_data.evaluation_run_id)


# ---------------------------------------------------------------------------
# Evaluation orchestration — merge + judgment + persist
# ---------------------------------------------------------------------------


def _execute_evaluation_runner(
    session: Any,
    run_data: EvaluationRunData,
    evaluation_instance: BaseEvaluationInstance,
    node_run_result_mapping_list: list[dict[str, NodeRunResult]],
    workflow_run_id_map: dict[int, str] | None = None,
) -> list[EvaluationItemResult]:
    """Evaluate all metrics, merge per-item, apply judgment, persist once.

    Ensures 1 test-data row = 1 EvaluationRunItem with all metrics combined.
    """
    results_by_index: dict[int, EvaluationItemResult] = {}

    # Phase 1: Default metrics — one batch per (metric, node) pair
    for default_metric in run_data.default_metrics:
        for node_info in default_metric.node_info_list:
            node_run_result_list: list[NodeRunResult] = []
            item_indices: list[int] = []
            for i, mapping in enumerate(node_run_result_mapping_list):
                node_result = mapping.get(node_info.node_id)
                if node_result is not None:
                    node_run_result_list.append(node_result)
                    item_indices.append(i)

            if not node_run_result_list:
                continue

            runner = _create_runner(EvaluationCategory(node_info.type), evaluation_instance)
            try:
                evaluated = runner.evaluate_metrics(
                    node_run_result_list=node_run_result_list,
                    default_metric=default_metric,
                    model_provider=run_data.evaluation_model_provider,
                    model_name=run_data.evaluation_model,
                    tenant_id=run_data.tenant_id,
                )
            except Exception:
                logger.exception(
                    "Failed metrics for %s on node %s", default_metric.metric, node_info.node_id
                )
                continue

            _stamp_and_merge(evaluated, item_indices, node_info, results_by_index)

    # Phase 2: Customized metrics
    if run_data.customized_metrics:
        try:
            customized_results = evaluation_instance.evaluate_with_customized_workflow(
                node_run_result_mapping_list=node_run_result_mapping_list,
                customized_metrics=run_data.customized_metrics,
                tenant_id=run_data.tenant_id,
            )
            for result in customized_results:
                _merge_result(results_by_index, result.index, result)
        except Exception:
            logger.exception("Failed customized metrics for run %s", run_data.evaluation_run_id)

    results = list(results_by_index.values())

    # Phase 3: Judgment
    if run_data.judgment_config:
        results = _apply_judgment(results, run_data.judgment_config)

    # Phase 4: Persist — one EvaluationRunItem per test-data row
    _persist_results(
        session, run_data.evaluation_run_id, results, run_data.input_list, workflow_run_id_map
    )

    return results


def _execute_retrieval_test(
    session: Any,
    evaluation_run: EvaluationRun,
    run_data: EvaluationRunData,
    evaluation_instance: BaseEvaluationInstance,
) -> list[EvaluationItemResult]:
    """Execute knowledge base retrieval for all items, then evaluate metrics.

    Unlike the workflow-based path, there are no workflow nodes to traverse.
    Hit testing is run directly for each dataset item and the results are fed
    straight into :class:`RetrievalEvaluationRunner`.
    """
    node_run_result_list = EvaluationService.execute_retrieval_test_targets(
        dataset_id=run_data.target_id,
        account_id=evaluation_run.created_by,
        input_list=run_data.input_list,
    )

    results_by_index: dict[int, EvaluationItemResult] = {}
    runner = RetrievalEvaluationRunner(evaluation_instance)

    for default_metric in run_data.default_metrics:
        try:
            evaluated = runner.evaluate_metrics(
                node_run_result_list=node_run_result_list,
                default_metric=default_metric,
                model_provider=run_data.evaluation_model_provider,
                model_name=run_data.evaluation_model,
                tenant_id=run_data.tenant_id,
            )
            item_indices = list(range(len(node_run_result_list)))
            _stamp_and_merge(evaluated, item_indices, None, results_by_index)
        except Exception:
            logger.exception("Failed retrieval metrics for run %s", run_data.evaluation_run_id)

    results = list(results_by_index.values())

    if run_data.judgment_config:
        results = _apply_judgment(results, run_data.judgment_config)

    _persist_results(session, run_data.evaluation_run_id, results, run_data.input_list)

    return results


# ---------------------------------------------------------------------------
# Helpers — merge, judgment, persist
# ---------------------------------------------------------------------------


def _stamp_and_merge(
    evaluated: list[EvaluationItemResult],
    item_indices: list[int],
    node_info: NodeInfo | None,
    results_by_index: dict[int, EvaluationItemResult],
) -> None:
    """Attach node_info to each metric and merge into results_by_index."""
    for result in evaluated:
        original_index = item_indices[result.index]
        if node_info is not None:
            for metric in result.metrics:
                metric.node_info = node_info
        _merge_result(results_by_index, original_index, result)


def _merge_result(
    results_by_index: dict[int, EvaluationItemResult],
    index: int,
    new_result: EvaluationItemResult,
) -> None:
    """Merge new metrics into an existing result for the same index."""
    existing = results_by_index.get(index)
    if existing:
        merged_metrics = existing.metrics + new_result.metrics
        actual = existing.actual_output or new_result.actual_output
        results_by_index[index] = existing.model_copy(
            update={"metrics": merged_metrics, "actual_output": actual}
        )
    else:
        results_by_index[index] = new_result.model_copy(update={"index": index})


def _apply_judgment(
    results: list[EvaluationItemResult],
    judgment_config: JudgmentConfig,
) -> list[EvaluationItemResult]:
    """Evaluate pass/fail judgment conditions on each result's metrics."""
    judged: list[EvaluationItemResult] = []
    for result in results:
        if result.error is not None or not result.metrics:
            judged.append(result)
            continue
        metric_values: dict[tuple[str, str], object] = {
            (m.node_info.node_id, m.name): m.value for m in result.metrics if m.node_info
        }
        judgment_result = JudgmentProcessor.evaluate(metric_values, judgment_config)
        judged.append(result.model_copy(update={"judgment": judgment_result}))
    return judged


def _persist_results(
    session: Any,
    evaluation_run_id: str,
    results: list[EvaluationItemResult],
    input_list: list[EvaluationDatasetInput],
    workflow_run_id_map: dict[int, str] | None = None,
) -> None:
    """Persist evaluation results — one EvaluationRunItem per test-data row."""
    dataset_map = {item.index: item for item in input_list}
    wf_map = workflow_run_id_map or {}

    for result in results:
        item_input = dataset_map.get(result.index)
        run_item = EvaluationRunItem(
            evaluation_run_id=evaluation_run_id,
            workflow_run_id=wf_map.get(result.index),
            item_index=result.index,
            inputs=json.dumps(item_input.inputs) if item_input else None,
            expected_output=item_input.expected_output if item_input else None,
            actual_output=result.actual_output,
            metrics=json.dumps([m.model_dump() for m in result.metrics]) if result.metrics else None,
            judgment=json.dumps(result.judgment.model_dump()) if result.judgment else None,
            metadata_json=json.dumps(result.metadata) if result.metadata else None,
            error=result.error,
            overall_score=getattr(result, "overall_score", None),
        )
        session.add(run_item)

    session.commit()


def _create_runner(
    category: EvaluationCategory,
    evaluation_instance: BaseEvaluationInstance,
) -> BaseEvaluationRunner:
    """Create the appropriate runner for the evaluation category."""
    match category:
        case EvaluationCategory.LLM:
            return LLMEvaluationRunner(evaluation_instance)
        case EvaluationCategory.RETRIEVAL | EvaluationCategory.KNOWLEDGE_BASE:
            return RetrievalEvaluationRunner(evaluation_instance)
        case EvaluationCategory.AGENT:
            return AgentEvaluationRunner(evaluation_instance)
        case EvaluationCategory.WORKFLOW:
            return WorkflowEvaluationRunner(evaluation_instance)
        case EvaluationCategory.SNIPPET:
            return SnippetEvaluationRunner(evaluation_instance)
        case _:
            raise ValueError(f"Unknown evaluation category: {category}")


# ---------------------------------------------------------------------------
# Status / summary / XLSX / storage helpers (unchanged logic)
# ---------------------------------------------------------------------------


def _mark_run_failed(session: Any, run_id: str, error: str) -> None:
    """Mark an evaluation run as failed."""
    try:
        evaluation_run = session.query(EvaluationRun).filter_by(id=run_id).first()
        if evaluation_run:
            evaluation_run.status = EvaluationRunStatus.FAILED
            evaluation_run.error = error[:2000]
            evaluation_run.completed_at = naive_utc_now()
            session.commit()
    except Exception:
        logger.exception("Failed to mark run %s as failed", run_id)


def _compute_metrics_summary(
    results: list[EvaluationItemResult],
    judgment_config: JudgmentConfig | None,
) -> dict[str, Any]:
    """Compute aggregate metric and judgment summaries for an evaluation run."""
    summary: dict[str, Any] = {}

    if judgment_config is not None and judgment_config.conditions:
        evaluated_results: list[EvaluationItemResult] = [
            result for result in results if result.error is None and result.metrics
        ]
        passed_items = sum(1 for result in evaluated_results if result.judgment.passed)
        evaluated_items = len(evaluated_results)
        summary["_judgment"] = {
            "enabled": True,
            "logical_operator": judgment_config.logical_operator,
            "configured_conditions": len(judgment_config.conditions),
            "evaluated_items": evaluated_items,
            "passed_items": passed_items,
            "failed_items": evaluated_items - passed_items,
            "pass_rate": passed_items / evaluated_items if evaluated_items else 0.0,
        }

    return summary


def _generate_result_xlsx(
    input_list: list[EvaluationDatasetInput],
    results: list[EvaluationItemResult],
) -> bytes:
    """Generate result XLSX with input data, actual output, metric scores, and judgment."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Evaluation Results")
    ws.title = "Evaluation Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Collect all metric names
    all_metric_names: list[str] = []
    for result in results:
        for metric in result.metrics:
            if metric.name not in all_metric_names:
                all_metric_names.append(metric.name)

    # Collect all input keys
    input_keys: list[str] = []
    for item in input_list:
        for key in item.inputs:
            if key not in input_keys:
                input_keys.append(key)

    has_judgment = any(bool(r.judgment.condition_results) for r in results)

    judgment_headers = ["judgment"] if has_judgment else []
    headers = (
        ["index"] + input_keys + ["expected_output", "actual_output"] + all_metric_names + judgment_headers + ["error"]
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.column_dimensions["A"].width = 10
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    result_by_index = {r.index: r for r in results}

    for row_idx, item in enumerate(input_list, start=2):
        result = result_by_index.get(item.index)

        col = 1
        ws.cell(row=row_idx, column=col, value=item.index).border = thin_border
        col += 1

        for key in input_keys:
            val = item.inputs.get(key, "")
            ws.cell(row=row_idx, column=col, value=str(val)).border = thin_border
            col += 1

        ws.cell(row=row_idx, column=col, value=item.expected_output or "").border = thin_border
        col += 1

        ws.cell(row=row_idx, column=col, value=result.actual_output if result else "").border = thin_border
        col += 1

        metric_scores = {m.name: m.value for m in result.metrics} if result else {}
        for metric_name in all_metric_names:
            score = metric_scores.get(metric_name)
            ws.cell(row=row_idx, column=col, value=score if score is not None else "").border = thin_border
            col += 1

        if has_judgment:
            if result and result.judgment.condition_results:
                judgment_value = "Pass" if result.judgment.passed else "Fail"
            else:
                judgment_value = ""
            ws.cell(row=row_idx, column=col, value=judgment_value).border = thin_border
            col += 1

        ws.cell(row=row_idx, column=col, value=result.error if result else "").border = thin_border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _store_result_file(
    tenant_id: str,
    run_id: str,
    xlsx_content: bytes,
    session: Any,
    created_by: str,
) -> str | None:
    """Store result XLSX file and return the UploadFile ID."""
    try:
        from extensions.ext_storage import storage
        from libs.uuid_utils import uuidv7

        filename = f"evaluation-result-{run_id[:8]}.xlsx"
        storage_key = f"evaluation_results/{tenant_id}/{str(uuidv7())}.xlsx"

        storage.save(storage_key, xlsx_content)

        upload_file: UploadFile = UploadFile(
            tenant_id=tenant_id,
            storage_type=dify_config.STORAGE_TYPE,
            key=storage_key,
            name=filename,
            size=len(xlsx_content),
            extension="xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            created_by_role=CreatorUserRole.ACCOUNT,
            created_by=created_by,
            created_at=naive_utc_now(),
            used=False,
        )
        session.add(upload_file)
        session.commit()
        return upload_file.id
    except Exception:
        session.rollback()
        logger.exception("Failed to store result file for run %s", run_id)
        return None
