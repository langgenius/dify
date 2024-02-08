import logging

from langchain.document_loaders.base import BaseLoader
from langchain.schema import Document
from openpyxl.reader.excel import load_workbook

logger = logging.getLogger(__name__)


class ExcelLoader(BaseLoader):
    """Load xlxs files.


    Args:
        file_path: Path to the file to load.
    """

    def __init__(
        self,
        file_path: str
    ):
        """Initialize with file path."""
        self._file_path = file_path

    def load(self) -> list[Document]:
        data = []
        keys = []
        wb = load_workbook(filename=self._file_path, read_only=True)
        # loop over all sheets
        for sheet in wb:
            if 'A1:A1' == sheet.calculate_dimension():
                sheet.reset_dimensions()
            for row in sheet.iter_rows(values_only=True):
                if all(v is None for v in row):
                    continue
                if keys == []:
                    keys = list(map(str, row))
                else:
                    row_dict = dict(zip(keys, list(map(str, row))))
                    row_dict = {k: v for k, v in row_dict.items() if v}
                    item = ''.join(f'{k}:{v};' for k, v in row_dict.items())
                    document = Document(page_content=item, metadata={'source': self._file_path})
                    data.append(document)

        return data
