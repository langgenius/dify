from pathlib import Path
import json
from typing import Dict
from openpyxl import load_workbook

from llama_index.readers.file.base_parser import BaseParser
from flask import current_app


class XLSXParser(BaseParser):
    """XLSX parser."""

    def _init_parser(self) -> Dict:
        """Init parser"""
        return {}

    def parse_file(self, file: Path, errors: str = "ignore") -> str:
        data = []
        keys = []
        with open(file, "r") as fp:
            wb = load_workbook(filename=file, read_only=True)
            # loop over all sheets
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    if all(v is None for v in row):
                        continue
                    if keys == []:
                        keys = list(map(str, row))
                    else:
                        data.append(json.dumps(dict(zip(keys, list(map(str, row)))), ensure_ascii=False))
        return '\n\n'.join(data)
