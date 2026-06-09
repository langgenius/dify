"""Excel document extractor used for RAG ingestion.

Supports cell hyperlinks for both `.xls` and `.xlsx`, and embedded worksheet images
for `.xlsx` files by converting them into markdown image links. Embedded images are
stored with deterministic keys derived from the source upload file and anchor cell so
retries can safely reuse the same assets.
"""

import hashlib
import logging
import mimetypes
import os
from typing import TypedDict, override

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import select

from configs import dify_config
from core.db.session_factory import session_factory
from core.rag.extractor.extractor_base import BaseExtractor
from core.rag.models.document import Document
from extensions.ext_storage import storage
from extensions.storage.storage_type import StorageType
from libs.datetime_utils import naive_utc_now
from models.enums import CreatorUserRole
from models.model import UploadFile

logger = logging.getLogger(__name__)


class Candidate(TypedDict):
    idx: int
    count: int
    map: dict[int, str]


class SheetImageCandidate(TypedDict):
    anchor: tuple[int, int]
    content_hash: str
    file_key: str
    image_bytes: bytes
    image_ext: str


class ExcelExtractor(BaseExtractor):
    """Load Excel files.

    Args:
        file_path: Path to the file to load.
    """

    _file_path: str
    _encoding: str | None
    _autodetect_encoding: bool
    _tenant_id: str | None
    _user_id: str | None
    _source_file_id: str | None

    def __init__(
        self,
        file_path: str,
        tenant_id: str | None = None,
        user_id: str | None = None,
        source_file_id: str | None = None,
        encoding: str | None = None,
        autodetect_encoding: bool = False,
    ):
        """Initialize with file path."""
        self._file_path = file_path
        self._tenant_id = tenant_id
        self._user_id = user_id
        self._source_file_id = source_file_id
        self._encoding = encoding
        self._autodetect_encoding = autodetect_encoding

    @override
    def extract(self) -> list[Document]:
        """Load from Excel file in xls or xlsx format using Pandas and openpyxl."""
        documents = []
        file_extension = os.path.splitext(self._file_path)[-1].lower()

        if file_extension == ".xlsx":
            # Worksheet drawing objects, including embedded images, are not available in read-only mode.
            wb = load_workbook(self._file_path, data_only=True)
            try:
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    header_row_idx, column_map, max_col_idx = self._find_header_and_columns(sheet)
                    if not column_map:
                        continue
                    start_row = header_row_idx + 1
                    sheet_image_map = self._extract_images_from_sheet(
                        sheet_name=sheet_name,
                        sheet=sheet,
                        valid_columns={column_idx + 1 for column_idx in column_map},
                        min_row=start_row,
                    )
                    for row in sheet.iter_rows(min_row=start_row, max_col=max_col_idx, values_only=False):
                        page_content = []
                        row_has_content = False
                        for col_idx, cell in enumerate(row):
                            value = cell.value
                            if col_idx in column_map:
                                col_name = column_map[col_idx]
                                if hasattr(cell, "hyperlink") and cell.hyperlink:
                                    target = getattr(cell.hyperlink, "target", None)
                                    if target:
                                        display_value = value if value is not None and str(value).strip() else target
                                        value = f"[{display_value}]({target})"
                                cell_row = getattr(cell, "row", None)
                                cell_column = getattr(cell, "column", None)
                                image_links = (
                                    sheet_image_map.get((cell_row, cell_column), [])
                                    if isinstance(cell_row, int) and isinstance(cell_column, int)
                                    else []
                                )
                                if value is None:
                                    value = ""
                                elif not isinstance(value, str):
                                    value = str(value)
                                if image_links:
                                    value = " ".join(filter(None, [value, " ".join(image_links)]))
                                value = value.strip()
                                if value:
                                    row_has_content = True
                                value = value.replace('"', '\\"')
                                page_content.append(f'"{col_name}":"{value}"')
                        if row_has_content and page_content:
                            documents.append(
                                Document(page_content=";".join(page_content), metadata={"source": self._file_path})
                            )
            finally:
                wb.close()

        elif file_extension == ".xls":
            excel_file = pd.ExcelFile(self._file_path, engine="xlrd")
            for excel_sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name=excel_sheet_name)
                df.dropna(how="all", inplace=True)

                for _, series_row in df.iterrows():
                    page_content = []
                    for k, v in series_row.items():
                        if pd.notna(v):
                            page_content.append(f'"{k}":"{v}"')
                    documents.append(
                        Document(page_content=";".join(page_content), metadata={"source": self._file_path})
                    )
        else:
            raise ValueError(f"Unsupported file extension: {file_extension}")

        return documents

    def _extract_images_from_sheet(
        self, sheet_name: str, sheet, valid_columns: set[int], min_row: int
    ) -> dict[tuple[int, int], list[str]]:
        """
        Extract embedded worksheet images and map them to their anchor cell.

        Images are stored with deterministic keys derived from the source upload file,
        sheet, anchor cell, and content hash so retried tasks can reuse the same
        UploadFile rows and storage objects.
        """
        if not self._tenant_id or not self._user_id or not self._source_file_id:
            return {}

        images = getattr(sheet, "_images", None) or []
        image_candidates: list[SheetImageCandidate] = []

        for image in images:
            marker = getattr(getattr(image, "anchor", None), "_from", None)
            row_idx = getattr(marker, "row", None)
            col_idx = getattr(marker, "col", None)
            if row_idx is None or col_idx is None:
                continue
            if row_idx + 1 < min_row or col_idx + 1 not in valid_columns:
                continue

            image_bytes = self._get_image_bytes(image)
            if not image_bytes:
                continue

            image_ext = self._get_image_extension(image)
            if not image_ext:
                continue

            anchor_row = row_idx + 1
            anchor_column = col_idx + 1
            content_hash = self._hash_image_bytes(image_bytes)
            image_candidates.append(
                {
                    "anchor": (anchor_row, anchor_column),
                    "content_hash": content_hash,
                    "file_key": self._build_image_file_key(
                        sheet_name=sheet_name,
                        anchor_row=anchor_row,
                        anchor_column=anchor_column,
                        content_hash=content_hash,
                        image_ext=image_ext,
                    ),
                    "image_bytes": image_bytes,
                    "image_ext": image_ext,
                }
            )

        if not image_candidates:
            return {}

        image_map: dict[tuple[int, int], list[str]] = {}
        base_url = dify_config.FILES_URL
        candidate_keys = sorted({candidate["file_key"] for candidate in image_candidates})

        with session_factory.create_session() as session:
            existing_upload_files = session.scalars(
                select(UploadFile).where(
                    UploadFile.tenant_id == self._tenant_id,
                    UploadFile.key.in_(candidate_keys),
                )
            ).all()
            upload_files_by_key = {upload_file.key: upload_file for upload_file in existing_upload_files}
            new_upload_files: list[UploadFile] = []

            for candidate in image_candidates:
                upload_file = upload_files_by_key.get(candidate["file_key"])
                if upload_file is None:
                    storage.save(candidate["file_key"], candidate["image_bytes"])
                    mime_type, _ = mimetypes.guess_type(candidate["file_key"])
                    upload_file = UploadFile(
                        tenant_id=self._tenant_id,
                        storage_type=StorageType(dify_config.STORAGE_TYPE),
                        key=candidate["file_key"],
                        name=candidate["file_key"],
                        size=len(candidate["image_bytes"]),
                        extension=candidate["image_ext"],
                        mime_type=mime_type or "",
                        created_by=self._user_id,
                        created_by_role=CreatorUserRole.ACCOUNT,
                        created_at=naive_utc_now(),
                        used=True,
                        used_by=self._user_id,
                        used_at=naive_utc_now(),
                        hash=candidate["content_hash"],
                    )
                    upload_files_by_key[candidate["file_key"]] = upload_file
                    new_upload_files.append(upload_file)

                image_map.setdefault(candidate["anchor"], []).append(
                    f"![image]({base_url}/files/{upload_file.id}/file-preview)"
                )

            if new_upload_files:
                session.add_all(new_upload_files)
                session.commit()

        return image_map

    @staticmethod
    def _hash_image_bytes(image_bytes: bytes) -> str:
        """Return a stable content hash for extracted image bytes."""
        return hashlib.sha256(image_bytes).hexdigest()

    def _build_image_file_key(
        self,
        *,
        sheet_name: str,
        anchor_row: int,
        anchor_column: int,
        content_hash: str,
        image_ext: str,
    ) -> str:
        """Build a deterministic storage key for an embedded worksheet image."""
        assert self._tenant_id is not None, "tenant_id is required for image extraction"
        assert self._source_file_id is not None, "source_file_id is required for image extraction"

        normalized_ext = image_ext.strip().lower()
        sheet_hash = hashlib.sha256(sheet_name.encode("utf-8")).hexdigest()[:16]
        return (
            f"image_files/{self._tenant_id}/{self._source_file_id}/"
            f"{sheet_hash}_r{anchor_row}_c{anchor_column}_{content_hash}.{normalized_ext}"
        )

    def _get_image_bytes(self, image) -> bytes | None:
        """Return embedded image bytes from an openpyxl image object."""
        data_loader = getattr(image, "_data", None)
        if not callable(data_loader):
            return None

        try:
            data = data_loader()
            if isinstance(data, bytes):
                return data
            if isinstance(data, bytearray):
                return bytes(data)
            logger.warning("Unexpected embedded image payload type: %s", type(data).__name__)
            return None
        except Exception:
            logger.warning("Failed to read embedded image bytes from Excel sheet", exc_info=True)
            return None

    def _get_image_extension(self, image) -> str | None:
        """Resolve an image extension from openpyxl metadata."""
        image_format = getattr(image, "format", None)
        if isinstance(image_format, str) and image_format.strip():
            return image_format.strip().lower()

        image_path = getattr(image, "path", None)
        if isinstance(image_path, str):
            _, extension = os.path.splitext(image_path)
            if extension:
                return extension.lstrip(".").lower()

        return None

    def _find_header_and_columns(self, sheet, scan_rows=10) -> tuple[int, dict[int, str], int]:
        """
        Scan first N rows to find the most likely header row.
        Returns:
            header_row_idx: 1-based index of the header row
            column_map: Dict mapping 0-based column index to column name
            max_col_idx: 1-based index of the last valid column (for iter_rows boundary)
        """
        # Store potential candidates: (row_index, non_empty_count, column_map)
        candidates: list[Candidate] = []

        # Limit scan to avoid performance issues on huge files
        # We iterate manually to control the read scope
        for current_row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=scan_rows, values_only=True), start=1):
            # Filter out empty cells and build a temp map for this row
            # col_idx is 0-based
            row_map = {}
            for col_idx, cell_value in enumerate(row):
                if cell_value is not None and str(cell_value).strip():
                    row_map[col_idx] = str(cell_value).strip().replace('"', '\\"')

            if not row_map:
                continue

            non_empty_count = len(row_map)

            # Header selection heuristic (implemented):
            # - Prefer the first row with at least 2 non-empty columns.
            # - Fallback: choose the row with the most non-empty columns
            #   (tie-breaker: smaller row index).
            candidates.append({"idx": current_row_idx, "count": non_empty_count, "map": row_map})

        if not candidates:
            return 0, {}, 0

        # Choose the best candidate header row.

        best_candidate: Candidate | None = None

        # Strategy: prefer the first row with >= 2 non-empty columns; otherwise fallback.

        for cand in candidates:
            if cand["count"] >= 2:
                best_candidate = cand
                break

        # Fallback: if no row has >= 2 columns, or all have 1, just take the one with max columns
        if not best_candidate:
            # Sort by count desc, then index asc
            candidates.sort(key=lambda x: (-x["count"], x["idx"]))
            best_candidate = candidates[0]

        # Determine max_col_idx (1-based for openpyxl)
        # It is the index of the last valid column in our map + 1
        max_col_idx = max(best_candidate["map"].keys()) + 1

        return best_candidate["idx"], best_candidate["map"], max_col_idx
